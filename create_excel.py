from openpyxl import load_workbook
import datetime
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,Color
import config
import re
import copy
import os
import modules.template_mgmt.model_api as ma
ma_obj =  ma.model_api()
import html_entity_to_single_char as html_entity
html_entity_obj = html_entity.html_entity_to_single_char()

class excel:

    def __init__(self):
        pass

    def get_clean_val(self, value_org):
        value = str(value_org)
        value = ''.join(value.split('%'))
        value = ''.join(value.split(','))
        value = ''.join(value.split('$'))
        try:
            value = int(value)
            return value
        except:
            try:
                value = float(value)
                return value
            except:pass
        try:
            return html_entity_obj.convert(value)
        except:pass 
        return value_org
        

    def find_new_row(self, r, ph_rows):
        new_rows = []
        for dd in ph_rows:
            if r > dd:
                new_rows.append(dd)
        return (r - len(new_rows))
    
    def create_excel(self, data, ws1, extra_info, consider_column= [], merge_cell_info={}, ph_flag= 'N', ph_rows= []):
        temp_c = 0 
        column_map_dic = {}
        data, isdata_available    =self.filter_data(data, merge_cell_info, conside_column, extra_info, ph_flag)
        for dd in data:
            r,c, value, taxonomy,formular_str,cell_alph = dd 
            col = c
            crow = row
            if 0:
                cc_ty = (r, c)
                if cc_ty in  merge_cell_info:
                    for mcell in merge_cell_info[cc_ty]:
                        if mcell == cc_ty:continue
                        data.append([mcell[0], mcell[1], value, '', '', ''])
                if r in ph_rows:continue
                if ph_rows:
                    r = self.find_new_row(r, ph_rows)
                if consider_column and (c not in consider_column):
                    continue
                if consider_column and (c in consider_column):
                    #print 'ccc', consider_column
                    if c not in column_map_dic:
                        temp_c = temp_c + 1    
                        column_map_dic[c] = temp_c
                    c = column_map_dic[c]
                else:
                    c = c + 1
            r = r + 1
            org_val = eval(value)
            cval = eval(value).get('value', '')
            #if org_val.get('Error_flg', []):
            #    return ['N', [org_val['Error_flg'], crow, col]]
            #if ph_flag == 'Y' and org_val.get('cl_t') == 'PH':continue
            cval = self.get_clean_val(cval)
            #print r,c , cval, org_val
            ws1.cell(column=c, row=r).value = cval
            if org_val.get('background-color',''):
                h = org_val['background-color']
                color  = h.replace('#','')
                ws1.cell(column=c, row=r).fill = PatternFill(fgColor= color, fill_type = "solid")
            if org_val.get('color',''):
                h = org_val['color']
                h = h.replace('#','')
                color = h
                #color= Color('{}'.format( tuple(int(h[i:i+2], 16) for i in (0, 2, 4)) ))
                #print color
                ws1.cell(column=c, row=r).font = Font(color = color)
            if org_val.get('text-align', ''):
                #print 'rtrtrt', org_val['text-align']
                ws1.cell(column=c, row=r).alignment = Alignment(horizontal=org_val['text-align'])
        if extra_info.get('mergedCells', []) and ph_flag != 'Y':
            for mr in extra_info['mergedCells']:
                #print 'mr', mr
                ws1.merge_cells(mr)
        return ['Y', []]

    def filter_data(self, data, merge_cell_info, consider_column, extra_info, ph_flag):
        temp_c = 0 
        column_map_dic = {}
        header_col_d  = {}
        rc_d    = {}
        col_config  = {}
        header_ph   = {}
        data_exists = {}
        isdata_available    = 'N'
        #print consider_column
        val_row_d   = {}
        for dd in data:
            r,c, value, taxonomy,formular_str,cell_alph = dd 
            tmpcels    = [c]
            
            if (r, c) in  merge_cell_info:
                tmpcels    = map(lambda x:x[1], merge_cell_info[(r, c)])
            org_val = {}
            if value: 
                org_val =  eval(value)
                if r == 0 and org_val.get('c_type') != 'H': continue
            #print [r, c, org_val.get('value'), tmpcels]
            for c in tmpcels:
                if consider_column and (c not in consider_column):
                    continue
                rc_d.setdefault(r,  {})[c]  = [r,c, value, taxonomy,formular_str,cell_alph]
                if value: 
                    org_val =  eval(value)
                    if org_val.get('value'):
                        #print [org_val['value'], org_val.get('c_type'),  org_val.get('cl_t')]
                        if org_val.get('derived') != 'Y':
                            val_row_d[r]    = 1
                        data_exists[(r, c)] = org_val
                        if org_val.get('c_type') == 'H': 
                            header_ph[r]   = 'H'
                        elif org_val.get('cl_t') in {'PH':1, 'G':1}:
                            header_ph[r]   = 'PH'
                        pval_d  = org_val.get('c_c', {}).get('PossibleValue', {})
                        if org_val.get('c_c'):
                            col_config[c]   = org_val['c_c']
                        if pval_d and  org_val['value'] in extra_info.get('LOOKUP_MAP', {}):
                            #print org_val['value'], pval_d, extra_info.get('LOOKUP_MAP', {})
                            if pval_d.get('type') == 'LV' and pval_d.get('LV', '') in extra_info.get('LOOKUP_MAP', {})[org_val['value']]:
                                header_col_d[c]  =  extra_info['LOOKUP_MAP'][org_val['value']][pval_d['LV']]
        valid_rows  = {}
        rs  = rc_d.keys()
        rs.sort()
        n_data  = []
        rcnt    = 0
        for r in rs:
            cols    = rc_d[r].keys()
            cols.sort()
            f_value = 0
            if header_ph.get(r, '') == 'PH':
                if ph_flag == 'Y':
                    continue
                f_value   = 1
            elif header_ph.get(r, '') == 'H':
                f_value = 1
            elif r not in val_row_d:continue
            else:
                for c in cols:
                    if (r,c) not in data_exists:continue
                    if col_config.get(c, {}).get('IsMetadata') != 'Yes':
                        f_value = 1
                        isdata_available    = 'Y'
                        #break
                    cval = data_exists[(r,c)]['value']
                    map_val = header_col_d.get(c, {}).get(cval)
                    if map_val:
                        cval =  map_val
                        data_exists[(r, c)]['value'] = cval
                        rc_d[r][c][2]   = data_exists[(r, c)]
            if f_value == 1:
                cnt = 0
                for c in cols:
                    tmpdd   = copy.deepcopy(rc_d[r][c])
                    tmpdd[0]   = rcnt
                    tmpdd[1]   = cnt
                    tmpdd[2]   = str(tmpdd[2])
                    n_data.append(tmpdd)
                    cnt += 1
                rcnt    += 1
                
        data    = n_data
        return data, isdata_available

    def create_txt_file_with_comma_sep(self, data,des, sepr, conside_column, merge_cell_info, ph_rows, extra_info, ph_flag):
        data, isdata_available    =self.filter_data(data, merge_cell_info, conside_column, extra_info, ph_flag)
        if isdata_available == 'N':
            #print 'isdata_available ', [isdata_available]
            return isdata_available
        #print 'merge_cell_info', merge_cell_info
        if not sepr:
            sepr = ','
        data_map = {}
        max_r = 0
        max_col = 0
        for dd in data:
            r,c, value, taxonomy,formular_str,cell_alph = dd 
            print '00000', r,c , value
            data_map['%s_%s'%(r, c)] = dd
            if r > max_r:
                max_r = r
            if c > max_col:
                max_col = c
        sheet_data = []
        #print max_r, max_col
        for r in range(max_r+1):
            row_data  = []
            for c in range(max_col+1):
                key = "%s_%s"%(r,c)
                cc_ty = (r, c)
                if 0:
                    if key in data_map:
                        if cc_ty in  merge_cell_info:
                            for mcell in merge_cell_info[cc_ty]:
                                if mcell == cc_ty:continue
                                #print 'key', mcell, cc_ty
                                data_map['%s_%s'%mcell] = [mcell[0], mcell[1], data_map[key][2], '', '']
                    if r in ph_rows:continue 
                    if conside_column and (c not in conside_column):
                        #print c
                        continue
                key = "%s_%s"%(r,c)
                value = ''
                if key in data_map:
                    cell_value = eval(data_map[key][2])
                    value = eval(data_map[key][2]).get('value', '')
                value = self.get_clean_val(value)
                row_data.append(str(value))
            sepr = str(sepr)
            if sepr == ',':
                rval = sepr.join(map(lambda x:'"'+x+'"',row_data))
            else:
                rval = sepr.join(row_data)
            #print 'rvallllll', rval, sepr
            sheet_data.append(rval)
        #print len(sheet_data), des
        fs_str = '\n'.join(sheet_data)
        f = open(des, "w")
        f.write(fs_str)
        f.close()
        return isdata_available #[isdata_available, []]
        
            

    def create_zip(self, files, des):
        import zipfile
        zip_file = zipfile.ZipFile(des, 'w')
        done_name   = {}
        for f in files:
            org_f = f.split('/')[-1]
            if org_f in done_name:
                done_name[org_f]    = done_name[org_f]+1
                num = done_name[org_f]
                org_f   = org_f.split('.')
                org_f[0]    = org_f[0]+'-'+str(num)
                org_f   = '.'.join(org_f)
            else:
                done_name[org_f]    = 0
            zip_file.write(f, org_f, compress_type=zipfile.ZIP_DEFLATED)
        zip_file.close()
        return des
                
            
    def get_mcell_info(self, extra):
        ex_info = eval(extra[0])
        if ex_info.get('mergedCells'):
            consider_cells = {}
            mcell_info = ma_obj.convert_merge_cells_v1(ex_info['mergedCells'])
            for cell, minfo in mcell_info.items():
                colspan , rowspan ,sr, er, sc, ec = minfo
                #print 'ce;;', cell, minfo
                for rr in range(sr -1, er):
                    for cc in range(sc-1, ec):
                        #if rr == sr and cc == sc:continue
                        #print '\t', rr, cc
                        consider_cells.setdefault((sr -1, sc-1), []).append((rr, cc))
            #print 'ce;;', consider_cells[(sr-1, sc-1)]
            return consider_cells
        else:
            return {}
            

    def read_data(self, path, template_id, sheet_id):
        path = ijson['path']
        import sqlite_api as sq
        sq_obj = sq.sqlite_api(path)
        info = sq_obj.read_sheet_data(template_id, sheet_id)
        return info

    def start_creation(self, info, sq_obj, header_name_sheets, unqick_header_name):
        files = []
        ex = 'Y'
        return_info = ''
        for ijson in info:
            dd = ijson
            if dd.get('ftype', '').lower() == 'text':
                data = []
                for sheet in ijson['s']:
                    print 'EEEEEEE', sheet
                    #if str(sheet[1]) != '12':continue
                    data = sq_obj.read_sheet_data(sheet[0], sheet[1])
                    extra_info = sq_obj.read_sheet_extra_info(sheet[0], sheet[1])
                    merge_cell_info = self.get_mcell_info(extra_info)  
                    temp_path = ijson['name']+".csv"
                    consider_d = ijson.get(sheet, {})
                    conside_column = consider_d.get('cols', []) 
                    sheet_headers = header_name_sheets.get(ijson['org_sh'], {})
                    '''for mdd, mcc in merge_cell_info.items():
                        if mdd[1] not in sheet_headers:continue
                        for cvv in mcc:
                            sheet_headers[cvv[1]] = sheet_headers[mdd[1]]
                    #print 'conside', sheet, temp_path, conside_column, ijson
                    temp_val = {}
                    for column in conside_column:
                        ch_name = header_name_sheets.get(ijson['org_sh'], {}).get(column, '')
                        #print '\t ccc', ch_name, column, header_name_sheets.get(ijson['org_sh'], {}), ijson['org_sh']
                        if ch_name and temp_val.get(ch_name, ''): 
                            ex = 'N'    
                            return_info = "In %s sheet %s column header repeating"%(ijson['org_sh'], ch_name)
                            break
                        else:
                            temp_val[ch_name] = 1 
                    hcount  = len(unqick_header_name.get(ijson['org_sh'], {}).keys()) 
                    print 'tttttt', hcount, conside_column, unqick_header_name.get(ijson['org_sh'], {})
                    #if hcount != len(conside_column):
                    #    ex = 'N'    
                    #    return_info =  "In %s sheet header count not matched"%(ijson['org_sh'])
                    #    break'''
                    extra_info  = extra_info[0]
                    if not extra_info:
                        extra_info = '{}'
                    extra_info  = eval(extra_info)
                    isdata_available    = self.create_txt_file_with_comma_sep(data, temp_path,ijson.get('sep', ','), conside_column, merge_cell_info,  consider_d.get('ph_rows', []), extra_info, consider_d.get('flg', 'N'))      
                    if isdata_available == 'Y':
                        files.append(temp_path)
                    #elif isdata_available == 'N':
                    #    return ['N', [], '%s, %s'%(ijson['org_sh'], ' '.join(Error_info))] 
            elif dd.get('ftype', '') == 'csv':
                workbook = Workbook()
                workbook.remove(workbook.active)
                for sheet in ijson['s']:
                    consider_d = ijson.get(sheet, {})
                    conside_column = consider_d.get('cols', []) 
                    #print 'conside', sheet, conside_column
                    sheet_clear = self.get_clean_sheet_name(sheet[2])
                    ws1 = workbook.create_sheet(sheet_clear)
                    data = sq_obj.read_sheet_data(sheet[0], sheet[1])
                    extra_info = sq_obj.read_sheet_extra_info(sheet[0], sheet[1])
                    merge_cell_info = self.get_mcell_info(extra_info)  
                    if not extra_info:
                        extra_info = '{}'
                    extra_info = eval(extra_info[0])
                    isdata_available, Error_info = self.create_excel(data, ws1, extra_info, conside_column, merge_cell_info, consider_d.get('flg', 'N'), consider_d.get('ph_rows', []))            
                    if isdata_available == 'N':
                        return ['N', [], '%s, %s'%(ijson['org_sh'], ' '.join(extra_info))] 
                workbook.save(ijson['name']+".csv")     
                files.append(ijson['name']+".csv")
            else:
                workbook = Workbook()
                workbook.remove(workbook.active)
                for sheet in ijson['s']:
                    consider_d = ijson.get(sheet, {})
                    conside_column = consider_d.get('cols', []) 
                    print 'conside', sheet, conside_column
                    sheet_clear = self.get_clean_sheet_name(sheet[2])
                    ws1 = workbook.create_sheet(sheet_clear)
                    data = sq_obj.read_sheet_data(sheet[0], sheet[1])
                    extra_info = sq_obj.read_sheet_extra_info(sheet[0], sheet[1])
                    merge_cell_info = self.get_mcell_info(extra_info)  
                    if not extra_info:
                        extra_info = '{}'
                    extra_info = eval(extra_info[0])
                    fggg, fginfo = self.create_excel(data, ws1, extra_info, conside_column, merge_cell_info, consider_d.get('flg', 'N'), consider_d.get('ph_rows', []))            
                    #if fggg == 'N': 
                    #    return ['N', [], '%s, %s'%(ijson['org_sh'], ' '.join(fginfo))] 
                workbook.save(ijson['name']+".xlsx")     
                files.append(ijson['name']+".xlsx")
        return [ ex, files, return_info]
       
    def get_datetime_format(self, ttype, date_v=None):
        if date_v:
            now = date_v
        else:
            now = datetime.datetime.now()
        dd = {
            'DDMMYYYY': '%d%m%Y',
            'MM/DD/YYYY': '%m/%d/%Y',
            'YYYYDDMM': '%Y%d%m',
            'YYYYMMDD': '%Y%m%d',
            'YYYYDDMM_HH_MM': '%Y%d%m_%H_%M',
            'DDMMYYYY_HH_MM': '%d%m%Y_%H_%M',
        } 
        #print (ttype, dd.get(ttype))
        vt = now.strftime(dd.get(ttype, '%d%m%Y'))
        return vt

    def zipname_formatting(self, ijson, name):
        ftype = ijson['ftype']
        if ftype == 'WITH-COMPANY-NAME':
            return name
        elif ftype == 'WITH-DATE':
            fformat = ijson['format']
            return name+"_"+self.get_datetime_format(fformat)
        elif ftype == 'WITH-DATE-TIME':
            fformat = ijson['format']
            return name+"_"+self.get_datetime_format(fformat)
        
    def fname_formatting(self, ijson, name):
        ftype = ijson['ftype']
        #print 'ftypw', ftype, ijson
        if ftype == 'AS-IS-IT':
            return name
        elif ftype == 'WITH-COMPANY_NAME':
            return name+"_"+ijson['cmp']
        elif ftype == 'WITH-DATE':
            fformat = ijson['format']
            if 'ph_date' in ijson and ijson['ph_date'] :
                return name+"_"+self.get_datetime_format(fformat, ijson['ph_date'])
            return name+"_"+self.get_datetime_format(fformat)
        elif ftype == 'WITH-DATE-TIME':
            fformat = ijson['format']
            if ijson['ph_date']:
                return name+"_"+self.get_datetime_format(fformat, ijson['ph_date'])
            return name+"_"+self.get_datetime_format(fformat)
            

    def fname_construction(self, ijson, extra):
        cmp_name = ijson.get('cmp', '')
        dt_sp = ijson.get('dt_delim')
        nn = datetime.datetime.now()
        if ijson.get('fc', '') == 'wcdt':
            dt_str = '%d'+dt_sp+'%m'+dt_sp+'%Y'+dt_sp+'%H'+dt_sp+'%M'+dt_sp+'%S'
            dt = nn.strftime(dt_str)
            extra = extra +  dt
        #elif:
        #    pass
        return extra

    def get_group_info(self, sq_obj, template_id, sheet_id, col_config, sname, header_name_sheets, unqick_header_name):
        data = sq_obj.read_sheet_data(template_id, sheet_id)
        cl_tv = []
        temp_dic = {}
        lcols = {}
        all_cols = []
        non_ph_cols = {}
        ph_cols = {}
        ph_rows = {}
        for dd in data:
            r,c, value, taxonomy,formular_str,cell_alph = dd 
            if c not in all_cols:
                all_cols.append(c)
            value  = eval(value)
            cl_t = value.get('cl_t', '')
            ctype = value.get('c_type', '')
            if cl_t.lower() == 'l':
                lcols[c] = 1 #.append(c)
            if value.get('value', ''):
                if cl_t.lower() == 'ph':
                    ph_cols[c] = 1
                    ph_rows[r] = 1
                else:
                    non_ph_cols[c] = 1
                if ctype == 'H' and r == 0:
                    header_name_sheets.setdefault(sname, {})[c] = value['value'] 
                    unqick_header_name.setdefault(sname, {})[c] = value['value']
            if cl_t and value.get('v', ''):
                temp_dic.setdefault(value['cl_t'], {}).setdefault(value['v'], {})[c] = 1          
        
        ll = list(set(non_ph_cols.keys()) - set(ph_cols))
        #print 'temooo', temp_dic
        for llll in ll:
            lcols[llll] = 1 
        if not temp_dic:
            cl_tv.append([sheet_id, sname, all_cols, 'N', ph_rows])
        else:
            for cl, cl_info in temp_dic.items():
                if cl not in col_config:continue
                for ph, phinfo in cl_info.items():
                    #print 'llcols', sheet_id, ph, lcols, phinfo
                    combine_cols = phinfo.keys() + lcols.keys()
                    combine_cols = list(set(combine_cols))
                    cl_tv.append([sheet_id, ph, combine_cols, 'Y', ph_rows])
        #sys.exit()
        #print '0000000000', sheet_id, temp_dic, cl_tv
        return cl_tv

    def get_clean_sheet_name(self, name):
        name = re.sub(r"[*/\?:,\[\]]", "", name)
        return name

    def read_meta_info(self, company_id):
        cid_path = config.Config.comp_path.format(company_id)
        db_path = os.path.join(cid_path, 'document_info.db')
        import sqlite_api as sq_c1
        sqc_obj = sq_c1.sqlite_api(db_path)    
        data = sqc_obj.read_meta_info() 
        map_info = {}
        for dd in data:
            period_type, period , mt = dd
            mt = eval(mt)
            map_info['%s%s'%(period_type, period)] = mt['Document To']
        return map_info
            
            
            
        
    def cl_create(self, ijson):
        import error_cell as er
        er_obj = er.excel()
        error_info = er_obj.cl_create(ijson)
        if error_info['message'] != 'done':
            return  [error_info]
        path = config.Config.mapping_path.format(ijson['company_id'], ijson['project_id'], ijson['template_id'])
        import sqlite_api as sq
        sq_obj = sq.sqlite_api(path)    
        template_id = ijson['template_id']
        import sqlite_api as sq
        sheets = sq_obj.read_sheets(template_id) 
        delim_map_d = {
                        'tab':'\t',
                        'comma':',',
                        }
        cmp_meta_info = self.read_meta_info(ijson['company_id'])
        #print 'cmp', cmp_meta_info
        import os
        fpath = config.Config.output.format(ijson['company_id'], ijson['project_id'], ijson['template_id'])
        cmd  = "mkdir -p '%s'"%(fpath)
        os.system(cmd)
        map_sheet_info = []
        cc = "/mnt/eMB_db/company_management/project_configuration/%s/config.db"%(ijson['project_id'])
        import sqlite_api as sq1
        sq1_obj = sq1.sqlite_api(cc)
        export_config, file_name_config, zip_file_name_config  = sq1_obj.read_output_config(ijson['template_id'])     
        export_config = eval(export_config)
        file_name_config = eval(file_name_config)
        zip_file_name_config = eval(zip_file_name_config)
        old_export_ctype = export_config.get('ctype', '')
        if 'col_config' in export_config and export_config['col_config']:
            export_config['ctype'] = 'column-group' 
        file_name_config['cmp'] = ijson['company_name']
        ctype = export_config.get('ctype', '').lower()
        #print ctype
        header_name_sheets = {}
        unqick_header_name = {}
        if ctype == 'all':
            temp = {}
            fname = ''
            for sheet in sheets:
                temp.setdefault('s', []).append((template_id, sheet[0], sheet[1]))
                fname = fname +'_'+sheet[1]
            name = self.fname_formatting(file_name_config, ijson['template_name'])
            ftype  = export_config.get('ftype', '')
            temp['ftype'] = ftype
            temp['org_sh'] =  sheet[1]
            temp['name'] = fpath + name
            temp['sep'] = delim_map_d.get(export_config.get('delim', '').lower(), '')
            temp['fpath'] = fpath
            map_sheet_info.append(temp)
        elif ctype == 'group':
            sheet_map = {}    
            for sheet in sheets:
                sheet_map[sheet[0]] = sheet[1]
            fname = ''
            for ginfo in ijson['ginfo']:
                temp = {}
                fname = ''
                for sid in ginfo['sids']:
                    sname = sheet_map[sid]
                    temp.setdefault('s', []).append((template_id, sid, sname))
                    fname = fname  + sname
                ftype  = export_config.get('ftype', '')
                temp['ftype'] = ftype
                name = self.fname_formatting(file_name_config, ijson['template_name'])
                temp['name'] = fpath  + name
                temp['org_sh'] =  sheet[1]
                temp['sep'] = export_config.get('delim', '')
                temp['fpath'] = fpath
                #print temp
                map_sheet_info.append(temp)
        elif ctype == 'individual':
            for sheet in sheets:
                temp = {}
                fname = sheet[1]
                name = self.fname_formatting(file_name_config, fname)
                temp.setdefault('s', []).append((template_id, sheet[0], sheet[1]))
                ftype  = export_config.get('ftype', '')
                temp['ftype'] = ftype
                temp['fpath'] = fpath
                name = self.fname_formatting(file_name_config, fname)
                temp['name'] =fpath + name
                temp['org_sh'] =  sheet[1]
                temp['sep'] = delim_map_d.get(export_config.get('delim', '').lower(), '')
                map_sheet_info.append(temp)
        elif ctype == 'column-group':
            fff = 1
            for sheet in sheets:
                #sid = sh['sid']
                #colums = sh['colums']  
                #name = sh['name']
                #if sheet[1] != 'deal':continue
                if ijson.get('ss','') and sheet[1] != ijson.get('ss',''):continue
                sep_info = self.get_group_info(sq_obj, template_id, sheet[0], export_config['col_config'], sheet[1], header_name_sheets, unqick_header_name)
                #print header_name_sheets
                for vff in sep_info:
                    temp = {}
                    sid, name, colums, ph_flg, ph_rows = vff
                    #print '9090', [sid, name, colums]
                    vtuple = (template_id, sid, name, fff)
                    fff = fff + 1
                    temp.setdefault('s', []).append(vtuple)
                    temp[vtuple] = {'cols': colums, 'name': name, 'flg': ph_flg, 'ph_rows': ph_rows, 'org_sh': sheet[1]}
                    ftype  = export_config.get('ftype', '')
                    temp['ftype'] = ftype
                    temp['org_sh'] =  sheet[1]
                    temp['fpath'] = fpath
                    temp['ct'] = 'Y'
                    if ph_flg == 'Y':
                        if cmp_meta_info.get(name, ''):
                            #print 'with datetime', cmp_meta_info[name], name
                            ph_date = datetime.datetime.strptime(cmp_meta_info[name], '%m/%d/%Y')
                            file_name_config['ph_date'] = ph_date
                            name = self.fname_formatting(file_name_config, sheet[1])
                            #print 'nameee', name
                        else:
                            name = sheet[1]+"_"+cmp_meta_info.get(name, name)
                        #name = self.fname_formatting(file_name_config, sheet[1]+"_"+name)
                        temp['ph_name']     = name
                    else:
                        name = self.fname_formatting(file_name_config, sheet[1])
                    temp['name'] =fpath + name
                    temp['sep'] = delim_map_d.get(export_config.get('delim', '').lower(), '')
                    #if old_export_ctype != 'all':
                    map_sheet_info.append(copy.deepcopy(temp))
                    #print 'mmmmmmm',temp 
                #sys.exit()
                #temp = {}
            #if old_export_ctype == 'all':
            #    map_sheet_info.append(temp)
                        
            #print '---------',map_sheet_info
        print '-------- Before creation' 
        ex, files, return_info = self.start_creation(map_sheet_info, sq_obj, header_name_sheets, unqick_header_name)
        if ex == 'N':
            return [{'message': return_info}]
        zzip_name = self.zipname_formatting(zip_file_name_config, ijson['company_name'])
        des = ijson.get('zip_name', fpath+zzip_name+'.zip')
        self.create_zip(files, des)
        download_path = config.Config.download_path.format(ijson['company_id'], ijson['project_id'], ijson['template_id'], zzip_name)
        return [{'message':'done', 'path': download_path, 'des':des}]

if __name__ == '__main__':
    obj = excel()
    ijson = {"cmd_id":126,"company_id":1053724,"template_id":4,"company_name":"PeaksCLO3,Ltd.","project_id":5,"template_name":"CLO Template","user":"demo", "PRINT": "Y", "ss":"transactions"}
    res =  obj.cl_create(ijson)
    print res

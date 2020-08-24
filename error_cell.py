from openpyxl import load_workbook
import datetime
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,Color
import config
import re
import copy
import os
import modules.template_mgmt.model_api as ma
ma_obj =  ma.model_api()
class excel:

    def __init__(self):
        pass

        
    def get_mcell_info(self, extra):
        ex_info = eval(extra[0])
        if ex_info.get('mergedCells'):
            consider_cells = {}
            mcell_info = ma_obj.convert_merge_cells_v1(ex_info['mergedCells'])
            for cell, minfo in mcell_info.items():
                colspan , rowspan ,sr, er, sc, ec = minfo
                #print 'ce;;', cell, minfo
                for rr in range(sr -1, er):
                    for cc in range(sc-1, ec):
                        #if rr == sr and cc == sc:continue
                        #print '\t', rr, cc
                        consider_cells.setdefault((sr -1, sc-1), []).append((rr, cc))
            #print 'ce;;', consider_cells[(sr-1, sc-1)]
            return consider_cells
        else:
            return {}
            
    def ph_name_wise_error(self, sheet, template_id, sq_obj): 
        ph_row = -1
        header_row = -1
        ph_col = {}
        header_col = {}
        ph_cell_missing = {}
        value_col = {}
        data = sq_obj.read_sheet_data_v1(template_id, sheet[0])
        extra_info = sq_obj.read_sheet_extra_info(template_id, sheet[0])
        merge_cell_info = self.get_mcell_info(extra_info)  
        Error_info = []
        cell_id_map = {}
        column_config = {}
        for dd in data:
            row_id, r,c, value, taxonomy,formular_str,cell_alph = dd 
            value  = eval(value)
            cl_t = value.get('cl_t', '').lower()
            ctype = value.get('c_type', '').lower()
            cell_value =  value.get('value', '')
            cell_id_map['%s_%s'%(r,c)] = row_id 
            if ctype == 'h':
                column_config[c] = value.get('c_c', {})
            if cl_t == 'ph' or  ctype == 'h':continue
            Error_flg = value.get('Error_flg', [])
            if sheet[1] == 'assets' and 'Data Type map value not Exisits' in Error_flg:continue 
            if Error_flg:
                Error_info.append([r,c, Error_flg])
        return [Error_info,column_config ]



    def full_sheet_validation(self, sheets, template_id, sq_obj, ijson):
        sheet_wise_dic = {}
        re_map = {}
        Error_message = ""
        Error_sheets = []
        for sheet in sheets:
            if ijson.get('sid', '') and sheet[0] != ijson['sid']:continue
            re_map[sheet[0]] = sheet[1]
            error_info, config_info = self.ph_name_wise_error(sheet,  template_id, sq_obj)
            if error_info:
                sheet_wise_dic[sheet[1]] = error_info 
                Error_sheets.append(sheet[1])
        if Error_sheets:
            Error_message = 'Errors avaiable sheets %s'%(','.join(Error_sheets))
        return sheet_wise_dic, Error_message
        
    def cl_create(self, ijson):
        path = config.Config.mapping_path.format(ijson['company_id'], ijson['project_id'], ijson['template_id'])
        import sqlite_api as sq
        sq_obj = sq.sqlite_api(path)    
        template_id = ijson['template_id']
        import sqlite_api as sq
        sheets = sq_obj.read_sheets(template_id) 
        error_data, Error_message = self.full_sheet_validation(sheets, template_id, sq_obj, ijson)
        if error_data:
            return {'message': Error_message, 'data': error_data}
        else:
            return {'message': 'done', 'data': []}

if __name__ == '__main__':
    obj = excel()
    ijson = {"cmd_id":126,"company_id":1053729,"template_id":4,"company_name":"KKRCLO21Ltd","project_id":5,"template_name":"CLO Template","user":"demo", "sid": 17} 
    res =  obj.cl_create(ijson)
    import json
    print json.dumps(res)

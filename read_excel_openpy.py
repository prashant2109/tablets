from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
import datetime, json

class read_excel:
    def read_excel(self, csv_file):
        f_d = {}
        wb = load_workbook(csv_file, data_only=False)
        sheets = wb.get_sheet_names()
        formula_info = {}
        order_of_sheet = []
        for sheet in sheets:
            ws = wb.get_sheet_by_name(sheet)
            sheetObj = ws
            for rowid, rowObjs in enumerate(sheetObj.rows):
                for colidx, cellObj in enumerate(rowObjs):
                    cell    = cellObj
                    formula = cell.value
                    if isinstance(formula, unicode):
                        formula = formula.encode('utf-8')
                    if isinstance(formula, str) and formula[0] == '=':
                        formula_info.setdefault(sheet, {})["%s_%s"%(rowid, colidx)] = formula
        wb = load_workbook(csv_file, data_only=True)
        sheets = wb.get_sheet_names()
        for sheet in sheets:
            ws = wb.get_sheet_by_name(sheet)
            data = []
            for row in range(1, ws.max_row+1):
                tmp_d = {'row': row-1, 'cols':{}}
                for col in range(1, ws.max_column+1):
                    cell = ws.cell(row=row, column=col)
                    fm_key = "%s_%s"%(row, col)
                    value = cell.value
                    if isinstance(cell.value, datetime.datetime):
                        value = value.strftime('%d-%b-%y')    
                    if isinstance(value, unicode):
                        value = value.encode('utf-8')
                    if isinstance(value, str):
                        value = ' '.join(value.strip().split())
                    if not value:
                        value   = ''
                    tmp_d['cols'][col -1] = {'v': value, 'fm': formula_info.get(sheet, {}).get(fm_key, '')}
                data.append(tmp_d)
            shh_name = sheet.strip()
            order_of_sheet.append(shh_name)
            f_d[shh_name] = data
        return [f_d, order_of_sheet]

if __name__ == '__main__':
    obj = read_excel() 
    res = obj.read_excel("SEM.xlsx") 
    print json.dumps(res)
